from flask import Blueprint, request, jsonify, send_file
from app.models import Estimate, EstimateDetail
from app import db
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break
from io import BytesIO
import os
from datetime import datetime
import subprocess

from app.models.database import connect_db


api_bp = Blueprint('api', __name__, url_prefix='/api')

# ✅ 見積一覧取得
@api_bp.route('/estimates', methods=['GET'])
def get_estimates():
    estimates = Estimate.query.all()
    result = []
    for e in estimates:
        result.append({
            "id": e.id,
            "project_name": e.project_name,
            "customer_name": e.customer_name,
            "total_cost": e.total_cost,
            "total_list_price": e.total_list_price,
            "quantity": e.quantity,
        })
    return jsonify(result)

# ✅ 明細データ取得
@api_bp.route('/estimate_details/<int:estimate_id>', methods=['GET'])
def get_estimate_details(estimate_id):
    details = EstimateDetail.query.filter_by(estimate_id=estimate_id).all()
    result = []
    for d in details:
        result.append({
            "id": d.id,
            "estimate_id": d.estimate_id,
            "item": d.item,
            "model": d.model,
            "quantity": d.quantity,
            "unit": d.unit,
            "cost_price": d.cost_price,
            "sale_price": d.sale_price,
            "cost_subtotal": d.cost_subtotal,
            "subtotal": d.subtotal,
        })
    return jsonify(result)

# ✅ 新規見積＋明細登録
@api_bp.route('/estimates', methods=['POST'])
def create_estimate():
    data = request.get_json()
    now = datetime.now()
    new_estimate = Estimate(
        project_name=data.get("project_name"),
        customer_name=data.get("customer_name"),
        total_cost=data.get("total_cost"),
        total_list_price=data.get("total_list_price"),
        quantity=data.get("quantity"),
        created_at=now.strftime("%Y-%m-%d"),
        year_month=now.strftime("%Y-%m")
    )
    db.session.add(new_estimate)
    db.session.commit()

    details = data.get("details", [])
    for d in details:
        detail = EstimateDetail(
            estimate_id=new_estimate.id,
            item=d.get("item"),
            model=d.get("model"),
            quantity=d.get("quantity"),
            unit=d.get("unit"),
            cost_price=d.get("cost_price"),
            sale_price=d.get("sale_price"),
            cost_subtotal=d.get("cost_subtotal"),
            subtotal=d.get("subtotal")
        )
        db.session.add(detail)

    db.session.commit()

    return jsonify({"message": "見積保存完了！", "id": new_estimate.id})

# ✅ 見積＋明細まとめて更新
@api_bp.route('/estimates/<int:estimate_id>', methods=['PUT'])
def update_estimate(estimate_id):
    data = request.get_json()
    estimate = Estimate.query.get(estimate_id)
    if not estimate:
        return jsonify({"message": "見積が存在しません"}), 404

    # 見積ヘッダー更新
    estimate.project_name = data.get("project_name")
    estimate.customer_name = data.get("customer_name")
    estimate.total_cost = data.get("total_cost")
    estimate.total_list_price = data.get("total_list_price")
    estimate.quantity = data.get("quantity")

    # 明細を一旦削除して新規登録
    EstimateDetail.query.filter_by(estimate_id=estimate_id).delete()
    for d in data.get("details", []):
        detail = EstimateDetail(
            estimate_id=estimate_id,
            item=d.get("item"),
            model=d.get("model"),
            quantity=d.get("quantity"),
            unit=d.get("unit"),
            cost_price=d.get("cost_price"),
            sale_price=d.get("sale_price"),
            cost_subtotal=d.get("cost_subtotal"),
            subtotal=d.get("subtotal")
        )
        db.session.add(detail)

    db.session.commit()

    return jsonify({"message": "見積＆明細 更新完了！"})

# ✅ 特定IDの見積取得
@api_bp.route('/estimates/<int:estimate_id>', methods=['GET'])
def get_estimate(estimate_id):
    estimate = Estimate.query.get(estimate_id)
    if not estimate:
        return jsonify({"message": "見積が存在しません"}), 404

    return jsonify({
        "id": estimate.id,
        "project_name": estimate.project_name,
        "customer_name": estimate.customer_name,
        "total_cost": estimate.total_cost,
        "total_list_price": estimate.total_list_price,
        "quantity": estimate.quantity
    })

## --- Excel出力＆PDF変換 ---
@api_bp.route('/export_excel/<int:estimate_id>', methods=['GET'])
def export_excel(estimate_id):
    try:
        # --- データ取得 ---
        estimate = Estimate.query.get(estimate_id)
        if not estimate:
            return jsonify({"error": "見積データが見つかりません"}), 404

        details = EstimateDetail.query.filter_by(estimate_id=estimate_id).all()

        # --- Excelテンプレートを開く ---
        template_path = os.path.join(os.path.dirname(__file__), '../../Excel原紙.xlsx')
        wb = load_workbook(template_path)
        ws = wb.active

        # --- 結合解除 ---
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))

        # --- ヘッダー情報書き込み ---
        ws["A4"] = estimate.customer_name or ""
        ws["F16"] = estimate.project_name or ""
        ws["J2"] = estimate.id
        ws["J3"] = datetime.now().strftime("%Y/%m/%d")

        # --- 明細データ書き込み ---
        start_row = 21
        for idx, d in enumerate(details):
            row = start_row + idx
            ws[f"A{row}"] = f"{d.item}（{d.model}）" if d.model else d.item
            ws[f"G{row}"] = d.quantity
            ws[f"H{row}"] = d.sale_price
            ws[f"I{row}"] = d.subtotal

        # --- 合計金額書き込み ---
        total = sum(d.subtotal for d in details)
        ws["I39"] = total
        ws["A12"] = total

        # --- 🔥 ファイル保存パス ---
        today_str = datetime.now().strftime("%Y%m%d")
        filename_base = f"{estimate.project_name}_{today_str}"
        excel_path = os.path.join(os.getcwd(), f"{filename_base}.xlsx")
        pdf_path = os.path.join(os.getcwd(), f"{filename_base}.pdf")

        # --- Excelファイル保存 ---
        wb.save(excel_path)
        print(f"✅ Excelファイル保存完了！保存先: {os.path.abspath(excel_path)}")

        # --- LibreOfficeでPDF変換 ---
        libreoffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        command = [
            libreoffice_path,
            "--headless",
            "--convert-to", "pdf",
            excel_path,
            "--outdir", os.getcwd()
        ]
        subprocess.run(command, check=True)

        # --- PDFをブラウザに送信 ---
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=f"{filename_base}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        print("エクセル出力エラー:", e)
        return jsonify({"error": str(e)}), 500
    
    
# ✅ テンプレート取得API
@api_bp.route("/templates/<int:id>", methods=["GET"])
def get_template(id):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM estimate_templates WHERE id = ?", (id,))
        template = cursor.fetchone()

        if not template:
            return jsonify({"error": "テンプレートが見つかりません"}), 404

        template_dict = {
            "id": template[0],
            "template_name": template[1],
            "project_name": template[2],
            "customer_name": template[3],
            "target_profit_rate": template[4],
            "category": template[5],
        }

        cursor.execute("SELECT * FROM estimate_template_details WHERE template_id = ?", (id,))
        rows = cursor.fetchall()

        details = []
        for row in rows:
            details.append({
                "item": row[2],
                "model": row[3],
                "quantity": row[4],
                "unit": row[5],
                "cost_price": row[6],
                "sale_price": row[7],
                "cost_subtotal": 0,  # 今は仮に0
                "subtotal": 0       # 今は仮に0
            })


        template_dict["details"] = details
        return jsonify(template_dict)

    except Exception as e:
        print("テンプレート取得エラー:", e)
        return jsonify({"error": "サーバーエラー"}), 500


@api_bp.route("/templates", methods=["POST"])
def save_template():
    try:
        data = request.get_json()

        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO estimate_templates (template_name, project_name, customer_name, target_profit_rate, category)
            VALUES (?, ?, ?, ?, ?)
        """, (
            data["template_name"],
            data.get("project_name", ""),
            data.get("customer_name", ""),
            data.get("target_profit_rate", 0),
            data.get("category", "")
        ))

        template_id = cursor.lastrowid

        for d in data.get("details", []):
            cursor.execute("""
                INSERT INTO estimate_template_details (template_id, item, model, quantity, unit, cost_price, sale_price, cost_subtotal, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                template_id,
                d.get("item", ""),
                d.get("model", ""),
                d.get("quantity", 0),
                d.get("unit", ""),
                d.get("cost_price", 0),
                d.get("sale_price", 0),
                d.get("cost_subtotal", 0),
                d.get("subtotal", 0)
            ))

        conn.commit()
        conn.close()

        return jsonify({"message": "✅ テンプレートを保存しました！"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        print("テンプレート保存エラー:", e)
        return jsonify({"error": "テンプレート保存に失敗しました"}), 500

@api_bp.route("/templates", methods=["GET"])
def get_templates():
    try:
        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM estimate_templates")
        rows = cursor.fetchall()

        result = []
        for row in rows:
            result.append({
                "id": row[0],
                "template_name": row[1],
                "project_name": row[2],
                "customer_name": row[3],
                "target_profit_rate": row[4],
                "category": row[5],
            })

        return jsonify(result)

    except Exception as e:
        print("テンプレート一覧取得エラー:", e)
        return jsonify({"error": "取得に失敗しました"}), 500
