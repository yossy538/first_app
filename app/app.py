# ==== 1. 標準ライブラリ ====
import os
import sys
import re
import subprocess
import sqlite3
from datetime import datetime

# ==== 2. 外部ライブラリ ====
import openpyxl
from flask import Flask, jsonify, send_file, render_template, Response, request

# ==== 3. 自作モジュール ====
# ==== 3. 自作モジュール ====
from app.models import db
from app.routes.main import main_bp
from app.routes.api import api_bp
from app.routes.upload_template_excel import upload_bp  # ← ここでまとめてimport！

# ==== 4. Flaskアプリ初期化 ====
app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///../database/estimates.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Blueprintを一気に登録
app.register_blueprint(main_bp)
app.register_blueprint(api_bp)
app.register_blueprint(upload_bp)  # ← これがちゃんと呼ばれるように！
 

# ==== 5. 定数・共通関数 ====
DB_PATH = os.path.join(os.path.dirname(__file__), "database", "estimates.db")

def connect_db():
    conn = sqlite3.connect(DB_PATH)
    return conn

def generate_filename(estimate_id, project_name, customer_name):
    today = datetime.now().strftime("%Y-%m-%d")  # ← 日付を 2025-04-15形式に！

    def clean_name(name):
        name = re.sub(r'[\\/:*?"<>|]', '_', name)
        return name.strip()

    clean_project_name = clean_name(project_name) or "未入力案件"
    clean_customer_name = clean_name(customer_name) or "未入力顧客"

    company_name = "株式会社サンクウ"

    # 🔥 ファイル名を作る
    filename = f"{estimate_id:04d}_{clean_customer_name}_{clean_project_name}_{today}_見積書_{company_name}.pdf"
    return filename



# ==== 6. ルーティング ====

@app.route("/")
def home():
    template_id = request.args.get("template_id", type=int)

    template_data = None
    if template_id:
        try:
            conn = connect_db()
            cursor = conn.cursor()

            # ヘッダー情報取得
            cursor.execute("""
                SELECT template_name, project_name, customer_name, target_profit_rate
                FROM estimate_templates
                WHERE id = ?
            """, (template_id,))
            header = cursor.fetchone()

            # 明細情報取得
            cursor.execute("""
                SELECT item, model, quantity, unit, cost_price, sale_price
                FROM estimate_template_details
                WHERE template_id = ?
            """, (template_id,))
            details = cursor.fetchall()
            conn.close()

            detail_list = [dict(
                item=d[0],
                model=d[1],
                quantity=d[2],
                unit=d[3],
                cost_price=d[4],
                sale_price=d[5],
                cost_subtotal=d[2] * d[4],
                subtotal=d[2] * d[5],
            ) for d in details]

            template_data = {
                "template_name": header[0],
                "project_name": header[1],
                "customer_name": header[2],
                "target_profit_rate": header[3],
                "details": detail_list
            }

        except Exception as e:
            print("テンプレート読み込みエラー:", e)

    return render_template("index.html", template_data=template_data)


    return render_template("index.html", estimates=estimates_list)

@app.route("/api/export_excel/<int:estimate_id>")
def export_excel(estimate_id):
    # --- データ取得 ---
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT project_name, customer_name FROM estimates WHERE id = ?", (estimate_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "指定された見積データが見つかりません"}), 404

    project_name, customer_name = row

    # --- ファイル名作成 ---
    base_filename = generate_filename(estimate_id, project_name, customer_name)  # ✅理想のファイル名
    target_pdf_path = f"./{base_filename}"
    temp_excel_path = "./temp_output.xlsx"   # Excel仮ファイル
    temp_pdf_path = "./temp_output.pdf"       # PDF仮ファイル（LibreOfficeが勝手に作る）


    # --- Excelファイル作成 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積書"
    ws["A1"] = "案件名"
    ws["B1"] = project_name
    ws["A2"] = "お客様名"
    ws["B2"] = customer_name
    wb.save(excel_path)

    print(f"✅ Excel保存完了: {excel_path}")

    # --- LibreOfficeでExcel → PDF変換 ---
    subprocess.run([
        "soffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", ".",  # 出力先フォルダ（カレントディレクトリ）
        excel_path
    ], check=True)

    print(f"✅ PDF変換完了: {pdf_path}")

    # --- PDFを返す ---
    return send_file(pdf_path, as_attachment=True)


def create_template_tables():
    conn = connect_db()
    cursor = conn.cursor()

    # テンプレート本体（ヘッダー）
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS estimate_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            template_name TEXT NOT NULL,
            project_name TEXT,
            customer_name TEXT,
            target_profit_rate REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # テンプレート明細
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS estimate_template_details (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            template_id INTEGER NOT NULL,
            item TEXT,
            model TEXT,
            quantity INTEGER,
            unit TEXT,
            cost_price REAL,
            sale_price REAL,
            FOREIGN KEY(template_id) REFERENCES estimate_templates(id)
        )
    """)

    conn.commit()
    conn.close()
    print("✅ テンプレート用テーブルを作成しました！")
    
@app.route("/api/templates", methods=["GET", "POST"])
def handle_templates():
    if request.method == "POST":
        try:
            data = request.get_json()

            conn = connect_db()
            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO estimate_templates (template_name, project_name, customer_name, target_profit_rate, category)
                VALUES (?, ?, ?, ?, ?)
            """, (
                data["template_name"],
                data.get("project_name", ""),
                data.get("customer_name", ""),
                data.get("target_profit_rate", 0),
                data.get("category", "")  # ← カテゴリも入れる場合
            ))

            template_id = cursor.lastrowid

            for detail in data["details"]:
                cursor.execute("""
                    INSERT INTO estimate_template_details (template_id, item, model, quantity, unit, cost_price, sale_price)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    template_id,
                    detail.get("item", ""),
                    detail.get("model", ""),
                    detail.get("quantity", 0),
                    detail.get("unit", ""),
                    detail.get("cost_price", 0),
                    detail.get("sale_price", 0)
                ))

            conn.commit()
            conn.close()

            return jsonify({"message": "✅ テンプレートを保存しました！"})

        except Exception as e:
            print("テンプレ保存エラー:", e)
            return jsonify({"error": str(e)}), 500

    elif request.method == "GET":
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, template_name, project_name, customer_name, category
                FROM estimate_templates
            """)
            rows = cursor.fetchall()
            conn.close()

            templates = []
            for row in rows:
                templates.append({
                    "id": row[0],
                    "template_name": row[1],
                    "project_name": row[2],
                    "customer_name": row[3],
                    "category": row[4]
                })

            return jsonify(templates)

        except Exception as e:
            print("テンプレ一覧取得エラー:", e)
            return jsonify({"error": str(e)}), 500



@app.route("/api/templates/<int:template_id>", methods=["GET"])
def get_template_by_id(template_id):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # ヘッダー情報取得
        cursor.execute("""
            SELECT template_name, project_name, customer_name, target_profit_rate
            FROM estimate_templates
            WHERE id = ?
        """, (template_id,))
        header = cursor.fetchone()

        if not header:
            return jsonify({"error": "テンプレートが見つかりません"}), 404

        # 明細データ取得
        cursor.execute("""
            SELECT item, model, quantity, unit, cost_price, sale_price
            FROM estimate_template_details
            WHERE template_id = ?
        """, (template_id,))
        details = cursor.fetchall()
        conn.close()

        detail_list = [dict(
            item=d[0],
            model=d[1],
            quantity=d[2],
            unit=d[3],
            cost_price=d[4],
            sale_price=d[5],
            cost_subtotal=d[2] * d[4],
            subtotal=d[2] * d[5]
        ) for d in details]

        result = {
            "template_name": header[0],
            "project_name": header[1],
            "customer_name": header[2],
            "target_profit_rate": header[3],
            "details": detail_list
        }

        return jsonify(result)

    except Exception as e:
        print("テンプレ読み込みエラー:", e)
        return jsonify({"error": str(e)}), 500


@app.route("/templates")
def templates_page():
    return render_template("templates.html")


@app.route("/api/templates/<int:template_id>", methods=["PUT"])
def update_template(template_id):
    try:
        data = request.get_json()

        conn = connect_db()
        cursor = conn.cursor()

        # テンプレートの存在チェック
        cursor.execute("SELECT id FROM estimate_templates WHERE id = ?", (template_id,))
        if not cursor.fetchone():
            return jsonify({"error": "指定されたテンプレートが見つかりません"}), 404

        # ヘッダー情報を更新（必要に応じて）
        cursor.execute("""
            UPDATE estimate_templates
            SET project_name = ?, customer_name = ?, target_profit_rate = ?
            WHERE id = ?
        """, (
            data.get("project_name", ""),
            data.get("customer_name", ""),
            data.get("target_profit_rate", 0),
            template_id
        ))

        # 明細を一度削除
        cursor.execute("DELETE FROM estimate_template_details WHERE template_id = ?", (template_id,))

        # 明細データを再挿入
        for detail in data["details"]:
            cursor.execute("""
                INSERT INTO estimate_template_details (template_id, item, model, quantity, unit, cost_price, sale_price)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                template_id,
                detail.get("item", ""),
                detail.get("model", ""),
                detail.get("quantity", 0),
                detail.get("unit", ""),
                detail.get("cost_price", 0),
                detail.get("sale_price", 0)
            ))

        conn.commit()
        conn.close()

        return jsonify({"message": "✅ テンプレートを更新しました！"})

    except Exception as e:
        print("テンプレート更新エラー:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/templates/<int:template_id>", methods=["DELETE"])
def delete_template(template_id):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # 明細削除 → ヘッダー削除
        cursor.execute("DELETE FROM estimate_template_details WHERE template_id = ?", (template_id,))
        cursor.execute("DELETE FROM estimate_templates WHERE id = ?", (template_id,))

        conn.commit()
        conn.close()

        return jsonify({"message": "テンプレートを削除しました！"})
    except Exception as e:
        print("テンプレート削除エラー:", e)
        return jsonify({"error": str(e)}), 500


@app.route("/api/estimates/<int:id>", methods=["DELETE"])
def delete_estimate(id):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # 明細データも先に削除（外部キー制約がある場合も）
        cursor.execute("DELETE FROM estimate_details WHERE estimate_id = ?", (id,))

        # 見積データ削除
        cursor.execute("DELETE FROM estimates WHERE id = ?", (id,))
        conn.commit()

        return jsonify({"message": "削除完了"})
    except Exception as e:
        print("削除エラー:", e)
        return jsonify({"error": "削除に失敗しました"}), 500



# ==== 7. 最後にアプリ起動 ====
if __name__ == "__main__":
    print(f"\U0001F4C1 使用しているDBパス: {DB_PATH}")
    create_template_tables()  # ← 💥これを一度だけ実行！
    app.run(debug=True, host="127.0.0.1", port=5002)
    

