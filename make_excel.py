from openpyxl import load_workbook

# ファイルパス
template_path = "./Excel原紙.xlsx"   # テンプレートパス
output_path = "./output_見積書.xlsx"  # 保存先パス

# データ（仮）
project_name = "サンプル案件名"
customer_name = "サンプル株式会社"

# 仮の明細データ
details = [
    {"item": "エアコン", "model": "AC-123", "quantity": 2, "unit_price": 50000, "subtotal": 100000},
    {"item": "室外機", "model": "OU-456", "quantity": 1, "unit_price": 80000, "subtotal": 80000},
    {"item": "配管セット", "model": "PIPE-789", "quantity": 3, "unit_price": 12000, "subtotal": 36000},
    # 必要に応じて追加していけるよ！！
]

# Excelテンプレートを開く
wb = load_workbook(template_path)

# シート選択（1枚目）
ws = wb.active

# ✅ 案件名・お客様名
ws["E15"] = project_name
ws["A4"] = customer_name

# ✅ 明細データを書き込み
start_row = 20  # 明細スタート行
for idx, detail in enumerate(details):
    row = start_row + idx
    ws[f"C{row}"] = f"{detail['item']}（{detail['model']}）"  # 品名＋型番
    ws[f"U{row}"] = detail["quantity"]                       # 数量
    ws[f"Y{row}"] = detail["unit_price"]                     # 単価
    ws[f"AE{row}"] = detail["subtotal"]                      # 金額

# ✅ 新しいファイルとして保存
wb.save(output_path)

print("✅ 案件名・お客様名・明細データを書き込みました！")
