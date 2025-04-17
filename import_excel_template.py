import openpyxl
import requests

# === ファイルパス ===
excel_path = "菱輝金型工業（株　AC8見積書20250123.xlsx"  # 正しいファイル名に直してね

# === Excel読み込み ===
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

# === ヘッダー情報（例）===
customer_name = ws["A4"].value or ""
project_name = ws["E14"].value or ""
target_profit_rate = 30
category = "AC工事"

# === 明細情報の抽出 ===
details = []
start_row = 19
while True:
    item = ws[f"C{start_row}"].value
    if item is None:
        break  # 空なら終了

    quantity = ws[f"U{start_row}"].value or 0
    unit = ws[f"W{start_row}"].value or ""
    cost_price = ws[f"Y{start_row}"].value or 0
    cost_subtotal = ws[f"AE{start_row}"].value or 0

    details.append({
        "item": str(item),
        "model": "",
        "quantity": quantity,
        "unit": unit,
        "cost_price": cost_price,
        "sale_price": 0,
        "cost_subtotal": cost_subtotal,
        "subtotal": 0
    })

    start_row += 1

# === APIに送信 ===
payload = {
    "template_name": "AC8_20250123",
    "project_name": project_name,
    "customer_name": customer_name,
    "target_profit_rate": target_profit_rate,
    "category": category,
    "details": details
}

res = requests.post("http://127.0.0.1:5002/api/templates", json=payload)

# === 結果表示 ===
try:
    print("📩 ステータス:", res.status_code)
    print("📦 結果:", res.json())
except Exception as e:
    print("⚠️ JSONデコード失敗:", e)
    print("レスポンス内容（テキスト）:", res.text)
